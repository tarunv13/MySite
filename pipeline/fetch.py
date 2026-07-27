#!/usr/bin/env python3
"""
Parallel Pipeline — data fetcher.

Pulls from the machine-readable EU sources and writes data/opportunities.json,
which index.html renders. Designed to run unattended in GitHub Actions.

WHAT THIS DOES AND DOES NOT DO
------------------------------
The Horizon Dashboard (dashboard.tech.ec.europa.eu) is a Qlik Sense app. It holds
no HTML — everything arrives over a websocket after JS boots — so it cannot be
fetched. Its underlying data is published separately, and that is what we use.

Two layers, and the distinction matters:

  LEADS   CORDIS bulk data. Who holds live grant money, when it ends, who the PI
          is. This is NOT a vacancy list. A live grant means a plausible need and
          a person worth writing to. Nothing more.

  CALLS   Funding & Tenders Portal (SEDIA). Real open calls with real deadlines.
          These are things you can actually apply to.

Actual PhD vacancies are announced by individual universities on their own job
boards. No single feed aggregates them reliably. Anyone claiming otherwise is
selling a scraper. The honest play is to work the leads layer, which is where
the advantage is anyway.

Usage:
    python fetch.py                  # full run
    python fetch.py --fixture X.csv  # test scoring against a local CSV
    python fetch.py --no-calls       # skip the SEDIA layer
"""

import argparse
import csv
import io
import json
import os
import re
import sys
import unicodedata
import zipfile
from datetime import datetime, timezone, date
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "opportunities.json"
UA = "ParallelPipeline/1.0 (+https://github.com/tarunv13/MySite) research-monitoring"
TIMEOUT = 180

# ── Sources ──────────────────────────────────────────────────────────────────
# CORDIS publishes monthly bulk snapshots. These URLs have been stable for years
# but are not contractual — if one 404s, check https://cordis.europa.eu/about/services
CORDIS = {
    "horizon": "https://cordis.europa.eu/data/cordis-HORIZONprojects-csv.zip",
    "h2020": "https://cordis.europa.eu/data/cordis-h2020projects-csv.zip",
}

# Funding & Tenders Portal search API. Public read, no EU Login needed.
SEDIA = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"
SEDIA_KEY = "SEDIA"  # public API key used by the portal's own front end

# ── Scoring vocabulary ───────────────────────────────────────────────────────
# Edit these. They encode one specific research profile; yours will differ.
TERMS = {
    "trade":       (5, r"wildlife trade|illegal wildlife|traffick|poach|CITES|bushmeat|ivory|wildlife crime"),
    "zoonotic":    (4, r"zoonot|spillover|One Health|emerging infectious|wildlife disease|host[- ]pathogen"),
    "culturomics": (5, r"culturomic|social media data|digital data|digital trace|crowdsourc|citizen science"),
    "language":    (3, r"multiling|natural language process|large language model|\bNLP\b|discourse|text mining|sentiment|multimodal"),
    "media":       (3, r"documentar|nature media|pro-environmental|behaviou?r change|environmental communication"),
    "biodiv":      (3, r"biodiversity|conservation|protected area|extinction|red list"),
    "spatial":     (2, r"spatial prioriti|conservation planning|remote sensing|land use change|urban green"),
}
MIN_FIT = 5
REQUIRE_ANY = {"biodiv", "culturomics", "trade", "zoonotic", "media"}

# ERC/Horizon panels worth keeping. Filters out the biology false positives that
# match "trade" in an ocean-microbiology abstract.
PANEL_OK = re.compile(r"\bSH[1-7]\b|\bLS8\b|\bPE10\b", re.I)

# Grants ending before this can't host a doctorate starting in 2027.
MIN_END = date(2028, 1, 1)

COMPILED = {k: (w, re.compile(p, re.I)) for k, (w, p) in TERMS.items()}


def log(msg):
    print(f"[{datetime.now(timezone.utc):%H:%M:%S}] {msg}", flush=True)


def get(url, data=None, headers=None):
    req = Request(url, data=data, headers={"User-Agent": UA, **(headers or {})})
    with urlopen(req, timeout=TIMEOUT) as r:
        return r.read()


def score(text, panel=""):
    """Return (fit, [matched tags]). Pure function — unit-testable."""
    blob = f"{text} {panel}"
    fit, tags = 0, []
    for key, (weight, rx) in COMPILED.items():
        if rx.search(blob):
            fit += weight
            tags.append(key)
    return fit, tags


def parse_date(s):
    if not s:
        return None
    s = str(s).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clean_org(s):
    return re.sub(r"\s*\[[^\]]*\]", "", str(s or "")).strip()


# ── CORDIS ───────────────────────────────────────────────────────────────────
def fetch_cordis(programme, url):
    log(f"CORDIS {programme}: downloading")
    try:
        raw = get(url)
    except (URLError, HTTPError, OSError) as e:
        log(f"  FAILED: {e}")
        return []
    log(f"  {len(raw)/1e6:.1f} MB")

    rows = []
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        name = next((n for n in z.namelist()
                     if "project" in n.lower() and n.lower().endswith(".csv")), None)
        if not name:
            log(f"  no project CSV in archive: {z.namelist()[:5]}")
            return []
        with z.open(name) as fh:
            text = io.TextIOWrapper(fh, encoding="utf-8", errors="replace")
            # CORDIS uses semicolons in some exports, commas in others. Sniff it.
            head = text.read(8192)
            text.seek(0)
            delim = ";" if head.count(";") > head.count(",") else ","
            for row in csv.DictReader(text, delimiter=delim):
                rows.append(row)
    log(f"  {len(rows)} projects parsed")
    return rows


def col(row, *names):
    """CORDIS column names drift between snapshots. Try several."""
    for n in names:
        for k in row:
            if k and k.strip().lower() == n.lower():
                return row[k]
    return ""


def process_cordis(rows, programme):
    out = []
    for r in rows:
        end = parse_date(col(r, "endDate", "End Date", "ecSignatureDate"))
        if not end or end < MIN_END:
            continue

        title = col(r, "title", "Project Title")
        abstract = col(r, "objective", "Abstract")
        panel = col(r, "panel", "Panel")

        fit, tags = score(f"{title} {abstract}", panel)
        if fit < MIN_FIT or not (set(tags) & REQUIRE_ANY):
            continue
        # Panel filter only where a panel is recorded (ERC rows have one)
        if panel and not PANEL_OK.search(panel):
            continue

        start = parse_date(col(r, "startDate", "Start Date"))
        out.append({
            "kind": "lead",
            "source": f"CORDIS {programme}",
            "id": col(r, "id", "Project Number", "projectID"),
            "acronym": col(r, "acronym", "Acronym"),
            "title": title.strip()[:300],
            "pi": col(r, "principalInvestigator", "Researcher(s)").strip(),
            "host": clean_org(col(r, "coordinator", "Host Institution(s)", "organisation"))[:160],
            "country": col(r, "coordinatorCountry", "Country", "country"),
            "grant_type": col(r, "fundingScheme", "Grant Type"),
            "panel": panel,
            "start": start.isoformat() if start else None,
            "end": end.isoformat(),
            "fresh": bool(start and (date.today() - start).days < 420),
            "fit": fit,
            "tags": tags,
            "url": col(r, "projectUrl", "CORDIS Link") or
                   f"https://cordis.europa.eu/project/id/{col(r, 'id', 'Project Number')}",
        })
    out.sort(key=lambda x: (-x["fit"], x["end"]), reverse=False)
    log(f"  {len(out)} scored above threshold")
    return out


# ── Funding & Tenders Portal ─────────────────────────────────────────────────
def fetch_calls(query="conservation OR biodiversity OR \"wildlife trade\" OR doctoral"):
    log("SEDIA: querying open calls")
    payload = json.dumps({
        "bool": {
            "must": [
                {"terms": {"type": ["1", "2", "8"]}},
                {"terms": {"status": ["31094501", "31094502"]}},  # forthcoming, open
            ]
        }
    }).encode()

    boundary = "----ParallelPipeline"
    parts = []
    for field, value in [("query", payload.decode()), ("languages", '["en"]')]:
        parts.append(f"--{boundary}\r\nContent-Disposition: form-data; name=\"{field}\"\r\n\r\n{value}\r\n")
    body = ("".join(parts) + f"--{boundary}--\r\n").encode()

    url = (f"{SEDIA}?apiKey={SEDIA_KEY}&text={query}"
           f"&pageSize=100&pageNumber=1&sortField=sortStatus&sortOrder=DESC")
    try:
        raw = get(url, data=body,
                  headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
        data = json.loads(raw)
    except Exception as e:
        log(f"  FAILED: {e} — continuing without calls layer")
        return []

    out = []
    for item in data.get("results", []) or []:
        md = item.get("metadata", {}) or {}

        def first(k):
            v = md.get(k)
            return v[0] if isinstance(v, list) and v else (v or "")

        title = first("title") or item.get("title", "")
        desc = re.sub(r"<[^>]+>", " ", str(first("description")))[:600]
        fit, tags = score(f"{title} {desc}")
        deadline = parse_date(str(first("deadlineDate"))[:10])

        out.append({
            "kind": "call",
            "source": "EU Funding & Tenders Portal",
            "id": first("identifier"),
            "title": title.strip()[:300],
            "programme": first("programmeDivision") or first("frameworkProgramme"),
            "status": first("status"),
            "deadline": deadline.isoformat() if deadline else None,
            "fit": fit,
            "tags": tags,
            "url": item.get("url", ""),
        })
    log(f"  {len(out)} calls retrieved")
    return out




# ── Roster cross-match ───────────────────────────────────────────────────────
ROSTER = ROOT / "data" / "people.json"


def deaccent(s):
    """Fold diacritics so 'Tromsø' matches 'Tromso' and 'Gagné' matches 'Gagne'.
    The ERC export and hand-typed roster entries disagree constantly on these."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return (s.replace("ø", "o").replace("Ø", "O")
             .replace("æ", "ae").replace("ß", "ss").replace("ð", "d").replace("þ", "th"))


def surname(full):
    """Last token that looks like a surname.

    Roster entries for multi-PI grants are written 'George Adamson (+ Gagné,
    Natarajan)', so anything parenthesised has to go before we look at tokens,
    or we end up searching CORDIS for 'Natarajan)'.
    """
    full = re.sub(r"\(.*?\)", " ", str(full))
    parts = [p for p in re.split(r"\s+", full.strip()) if len(p) > 1 and "." not in p]
    return parts[-1] if parts else full.strip()


def match_roster(all_rows):
    """Attach live grant records to roster entries.

    Matching is deliberately strict. A bare substring test is useless here:
    "Fink" matches "Finkenrath", and a UK grant by any "Jones" matches Kate Jones.
    So we require BOTH a word-boundary surname hit AND institution-word overlap,
    with a first-initial check as the only fallback. Under-matching is the correct
    failure mode — a missed grant costs you a lookup, a false one costs you an
    email that makes you look careless.
    """
    if not ROSTER.exists():
        log("no roster file — skipping cross-match")
        return []
    doc = json.loads(ROSTER.read_text(encoding="utf-8"))
    people = doc.get("people", [])

    STOP = {"university", "universite", "universiteit", "college", "institute",
            "school", "research", "national", "centre", "center", "technology",
            "science", "sciences", "london", "study", "studies", "department"}

    for p in people:
        sn = deaccent(surname(p["name"]))
        sn_rx = re.compile(rf"\b{re.escape(sn)}\b", re.I)
        initial = p["name"].strip()[0].lower()
        inst_words = {w.lower() for w in re.findall(r"[A-Za-z]{4,}", deaccent(p.get("inst", "")))} - STOP

        found = []
        for r in all_rows:
            pi = deaccent(col(r, "principalInvestigator", "Researcher(s)"))
            if not sn_rx.search(pi):
                continue
            # first-initial guard: "K. Jones" should not match "Trevor Jones"
            names = [n for n in pi.split(",")]
            if not any(sn_rx.search(n) and n.strip()[:1].lower() == initial for n in names):
                continue
            host = deaccent(clean_org(col(r, "coordinator", "Host Institution(s)", "organisation")))
            host_words = set(re.findall(r"[a-z]{4,}", host.lower())) - STOP
            if not (inst_words & host_words):
                continue
            end = parse_date(col(r, "endDate", "End Date"))
            found.append({
                "acronym": col(r, "acronym", "Acronym"),
                "title": col(r, "title", "Project Title")[:160],
                "end": end.isoformat() if end else None,
                "live": bool(end and end >= date.today()),
                "url": col(r, "projectUrl", "CORDIS Link"),
            })
        found.sort(key=lambda g: g.get("end") or "", reverse=True)
        p["grants"] = found[:4]
        p["live_grant"] = any(g["live"] for g in found)

    live = sum(1 for p in people if p.get("live_grant"))
    log(f"roster: {len(people)} people, {live} with a live grant")
    return people


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fixture", help="score a local CSV instead of downloading")
    ap.add_argument("--no-calls", action="store_true")
    ap.add_argument("--no-h2020", action="store_true")
    args = ap.parse_args()

    leads, calls, raw_all = [], [], []

    if args.fixture:
        log(f"fixture mode: {args.fixture}")
        if args.fixture.lower().endswith((".xlsx", ".xlsm")):
            try:
                from openpyxl import load_workbook
            except ImportError:
                sys.exit("xlsx fixture needs openpyxl: pip install openpyxl")
            wb = load_workbook(args.fixture, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            hdr = [str(h).strip() if h is not None else "" for h in next(it)]
            rows = [dict(zip(hdr, ["" if v is None else str(v) for v in r])) for r in it]
        else:
            with open(args.fixture, encoding="utf-8", errors="replace") as fh:
                rows = list(csv.DictReader(fh))
        raw_all = rows
        leads = process_cordis(rows, "fixture")
    else:
        for prog, url in CORDIS.items():
            if prog == "h2020" and args.no_h2020:
                continue
            rows = fetch_cordis(prog, url)
            raw_all += rows
            leads += process_cordis(rows, prog.upper())
        if not args.no_calls:
            calls = fetch_calls()

    # Dedupe by project id. The ERC dashboard export ships near-duplicate rows
    # (same project number, one copy with the acronym and one with "-"), so prefer
    # the richer record rather than whichever arrived first.
    def richness(x):
        return (x["fit"], x.get("acronym", "") not in ("", "-"),
                len(x.get("title", "")), len(x.get("pi", "")))

    seen = {}
    for l in leads:
        k = l.get("id") or l.get("acronym")
        if k not in seen or richness(l) > richness(seen[k]):
            seen[k] = l
    leads = sorted(seen.values(), key=lambda x: (-x["fit"], x["end"]))

    calls = [c for c in calls if c["fit"] >= 3 or c.get("deadline")]
    calls.sort(key=lambda c: (c["deadline"] or "9999", -c["fit"]))

    roster = match_roster(raw_all)

    doc = {
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "counts": {"leads": len(leads), "calls": len(calls), "people": len(roster)},
        "note": ("Leads are live grants, not vacancies — they identify who holds "
                 "money and is plausibly recruiting. Calls are applicable openings."),
        "leads": leads[:200],
        "calls": calls[:100],
        "roster": roster,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, indent=1, ensure_ascii=False), encoding="utf-8")
    log(f"wrote {OUT.relative_to(ROOT)} — {len(leads)} leads, {len(calls)} calls")

    if not leads and not calls and not roster:
        log("WARNING: empty result set — sources may have moved")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
